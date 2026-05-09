"""
Schritt 1 – Daten einlesen (01_load_data.py)

Liest alle 5 Einzelabschlüsse aus data/ ein (Sheet GuV + Bilanz),
normalisiert Positionsbezeichnungen und validiert Aktiva == Passiva.
"""

import json
import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)

ROOT = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"
CONFIG_DIR = ROOT / "config"

# GuV-Zeilen, die Kennzahlen (keine GuV-Positionen) sind – ab dieser Zeile ignorieren
_GUV_STOP_KEYWORDS = {"KENNZAHLEN", "LEGENDE"}


def _load_config() -> tuple[dict, dict]:
    ownership = json.loads((CONFIG_DIR / "ownership.json").read_text(encoding="utf-8"))
    pos_map = json.loads(
        (CONFIG_DIR / "position_mapping.json").read_text(encoding="utf-8")
    )
    return ownership, pos_map


def _build_reverse_map(pos_map: dict, sheet: str) -> dict[str, str]:
    """Liefert alias → kanonischer Name für einen Sheet-Typ (GuV oder Bilanz)."""
    mapping: dict[str, str] = {}
    for canonical, aliases in pos_map.get(sheet, {}).items():
        for alias in aliases:
            mapping[alias.strip().lower()] = canonical
    return mapping


def _normalize_position(raw: str, reverse_map: dict[str, str]) -> str:
    """Bereinigt Leerzeichen und ersetzt Aliase durch kanonische Namen."""
    clean = raw.strip()
    return reverse_map.get(clean.lower(), clean)


def _extract_currency(ws_guv) -> str:
    """Liest Währung aus der Spaltenkopfzeile (Zeile 3) der GuV."""
    for row in ws_guv.iter_rows(min_row=3, max_row=3, values_only=True):
        header = str(row[1]) if row[1] else ""
        match = re.search(r"\(T([A-Z]{3})\)", header)
        return match.group(1) if match else "EUR"
    return "EUR"


def _parse_guv(ws, reverse_map: dict[str, str]) -> pd.DataFrame:
    """Parst GuV-Sheet → DataFrame mit Index 'Position', Spalten '2024', '2023'."""
    rows = []
    stop = False
    for row in ws.iter_rows(min_row=4, values_only=True):
        pos_raw, val_2024, val_2023 = row[0], row[1], row[2]
        if pos_raw is None:
            continue
        pos_stripped = pos_raw.strip()
        if pos_stripped.upper() in _GUV_STOP_KEYWORDS:
            stop = True
        if stop:
            continue
        # Zeilen ohne Zahlenwerte sind Abschnittsüberschriften – überspringen
        if not isinstance(val_2024, (int, float)) and not isinstance(val_2023, (int, float)):
            continue
        pos_norm = _normalize_position(pos_stripped, reverse_map)
        rows.append(
            {
                "Position": pos_norm,
                "2024": float(val_2024) if isinstance(val_2024, (int, float)) else None,
                "2023": float(val_2023) if isinstance(val_2023, (int, float)) else None,
            }
        )
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    # Duplikate (z. B. IC-Unterpositionen mit gleichem Alias) aggregieren
    df = df.groupby("Position", sort=False)[["2024", "2023"]].sum()
    return df


def _parse_bilanz(ws, reverse_map: dict[str, str]) -> pd.DataFrame:
    """
    Parst das zweispaltige Bilanz-Sheet.
    Aktiva: Spalte A (idx 0), Werte C/D (idx 2/3)
    Passiva: Spalte E (idx 4), Werte G/H (idx 6/7)
    """
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        aktiva_pos = row[0]
        aktiva_24 = row[2]
        aktiva_23 = row[3]
        passiva_pos = row[4]
        passiva_24 = row[6]
        passiva_23 = row[7]

        if aktiva_pos and isinstance(aktiva_24, (int, float)):
            pos = _normalize_position(str(aktiva_pos), reverse_map)
            rows.append(
                {
                    "Position": pos,
                    "Seite": "Aktiva",
                    "2024": float(aktiva_24),
                    "2023": float(aktiva_23) if isinstance(aktiva_23, (int, float)) else None,
                }
            )
        if passiva_pos and isinstance(passiva_24, (int, float)):
            pos = _normalize_position(str(passiva_pos), reverse_map)
            rows.append(
                {
                    "Position": pos,
                    "Seite": "Passiva",
                    "2024": float(passiva_24),
                    "2023": float(passiva_23) if isinstance(passiva_23, (int, float)) else None,
                }
            )
    return pd.DataFrame(rows)


def _get_bilanzsumme(bilanz_df: pd.DataFrame, seite: str) -> Optional[float]:
    """Sucht BILANZSUMME AKTIVA bzw. PASSIVA im DataFrame."""
    keyword = f"BILANZSUMME {seite.upper()}"
    mask = (bilanz_df["Position"].str.upper() == keyword) & (bilanz_df["Seite"] == seite)
    werte = bilanz_df.loc[mask, "2024"]
    return float(werte.iloc[0]) if not werte.empty else None


def _validate_bilanz(bilanz_df: pd.DataFrame, gesellschaft: str, waehrung: str) -> bool:
    aktiva = _get_bilanzsumme(bilanz_df, "Aktiva")
    passiva = _get_bilanzsumme(bilanz_df, "Passiva")
    if aktiva is None or passiva is None:
        logger.warning("%s: Bilanzsummen konnten nicht ermittelt werden.", gesellschaft)
        return False
    diff = round(abs(aktiva - passiva), 2)
    if diff == 0:
        logger.info("%s: Bilanz ausgeglichen (%.0f T%s).", gesellschaft, aktiva, waehrung)
        return True
    logger.warning(
        "%s: Bilanz NICHT ausgeglichen – Aktiva=%.0f, Passiva=%.0f, Diff=%.0f T%s.",
        gesellschaft,
        aktiva,
        passiva,
        aktiva - passiva,
        waehrung,
    )
    return False


def load_all() -> dict[str, dict]:
    """
    Liest alle 5 Einzelabschlüsse ein.

    Rückgabe:
        {
          "muster_de": {
              "name":     str,
              "waehrung": str,          # Lokalwährung (EUR / CHF / PLN)
              "beteiligung": float,
              "minderheit":  float,
              "guv":    pd.DataFrame,   # Index=Position, Spalten 2024/2023
              "bilanz": pd.DataFrame,   # Spalten Position/Seite/2024/2023
              "bilanz_ok": bool,
          },
          ...
        }
    """
    ownership, pos_map = _load_config()
    rev_guv = _build_reverse_map(pos_map, "GuV")
    rev_bilanz = _build_reverse_map(pos_map, "Bilanz")

    result: dict[str, dict] = {}

    for key, meta in ownership.items():
        datei = DATA_DIR / f"{key}.xlsx"
        if not datei.exists():
            logger.error("Datei nicht gefunden: %s", datei)
            continue

        logger.info("Lade %s …", datei.name)
        wb = openpyxl.load_workbook(str(datei), read_only=True, data_only=True)

        if "GuV" not in wb.sheetnames or "Bilanz" not in wb.sheetnames:
            logger.error("%s: Sheets 'GuV' oder 'Bilanz' fehlen.", key)
            wb.close()
            continue

        waehrung = _extract_currency(wb["GuV"])
        guv_df = _parse_guv(wb["GuV"], rev_guv)
        bilanz_df = _parse_bilanz(wb["Bilanz"], rev_bilanz)
        wb.close()

        bilanz_ok = _validate_bilanz(bilanz_df, meta["name"], waehrung)

        result[key] = {
            "name": meta["name"],
            "land": meta["land"],
            "waehrung": waehrung,
            "beteiligung": meta["beteiligung"],
            "minderheit": meta["minderheit"],
            "guv": guv_df,
            "bilanz": bilanz_df,
            "bilanz_ok": bilanz_ok,
        }

    return result


def print_uebersicht(daten: dict[str, dict]) -> None:
    """Gibt eine formatierte Übersicht der Bilanzsummen je Gesellschaft aus."""
    sep = "─" * 92
    print()
    print(sep)
    print(
        f"{'Gesellschaft':<26} {'Land':<14} {'Währg':<5}"
        f" {'Aktiva 2024':>14} {'Passiva 2024':>14} {'Diff':>10}  {'Status'}"
    )
    print(sep)

    for key, d in daten.items():
        bilanz_df = d["bilanz"]
        aktiva = _get_bilanzsumme(bilanz_df, "Aktiva")
        passiva = _get_bilanzsumme(bilanz_df, "Passiva")
        diff = round(aktiva - passiva, 0) if aktiva and passiva else None
        status = "OK" if d["bilanz_ok"] else f"DIFF {diff:+,.0f}"
        w = d["waehrung"]
        print(
            f"{d['name']:<26} {d['land']:<14} {w:<5}"
            f" {aktiva:>12,.0f}  {passiva:>12,.0f}  {diff:>+10,.0f}  {status}"
        )

    print(sep)

    # GuV-Übersicht: Umsatz und Jahresüberschuss
    print()
    print(sep)
    print(
        f"{'Gesellschaft':<26} {'Währg':<5}"
        f" {'Umsatz 2024':>14} {'EBIT 2024':>12} {'JÜ 2024':>12}"
    )
    print(sep)
    for key, d in daten.items():
        guv = d["guv"]
        w = d["waehrung"]

        def _get(pos):
            try:
                return guv.loc[pos, "2024"]
            except KeyError:
                return None

        umsatz = _get("Gesamtumsatz")
        ebit = _get("EBIT")
        ju = _get("JAHRESÜBERSCHUSS")
        print(
            f"{d['name']:<26} {w:<5}"
            f" {(umsatz or 0):>12,.0f}  {(ebit or 0):>10,.0f}  {(ju or 0):>10,.0f}"
        )
    print(sep)
    print()


if __name__ == "__main__":
    daten = load_all()

    if not daten:
        logger.error("Keine Daten geladen – Abbruch.")
        raise SystemExit(1)

    logger.info("Einlesen abgeschlossen: %d Gesellschaften.", len(daten))
    print_uebersicht(daten)
