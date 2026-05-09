"""
Schritt 6 – Excel-Export (06_export_excel.py)

Erstellt output/Konzernabschluss_2024.xlsx mit 5 Sheets:
  1. Konzern-GuV       – Ertrag/Aufwand mit Vorjahresvergleich
  2. Konzern-Bilanz    – Aktiva / Passiva nebeneinander
  3. Kapitalflussrechnung – indirekte Methode
  4. Segmentbericht    – Kennzahlen je Gesellschaft / Region
  5. IC-Eliminierungen – Nachweis aller Buchungen

Farbcode (CLAUDE.md-Konvention):
  Blau  (#DDEEFF) – Eingaben / Rohdaten
  Schwarz auf Grau (#F2F2F2) – berechnete Positionen / Subtotals
  Grün  (#DDFFDD) – Verknüpfungen zwischen Sheets
  Rot   (#FFCCCC) – Warnhinweise / Differenzen
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)

ROOT = Path(__file__).parent.parent
OUTPUT_DIR = ROOT / "output"

# ---------------------------------------------------------------------------
# Farben & Stile
# ---------------------------------------------------------------------------

_BLUE   = PatternFill("solid", fgColor="DDEEFF")   # Eingaben
_GRAY   = PatternFill("solid", fgColor="F2F2F2")   # Subtotals / Formeln
_GREEN  = PatternFill("solid", fgColor="DDFFDD")   # Verknüpfungen
_RED    = PatternFill("solid", fgColor="FFCCCC")   # Warnungen
_HEADER = PatternFill("solid", fgColor="003366")   # Titelzeile
_SUBHDR = PatternFill("solid", fgColor="336699")   # Abschnittsköpfe

_BORDER_THIN = Border(
    bottom=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
)
_BORDER_THICK = Border(
    bottom=Side(style="medium", color="003366"),
    top=Side(style="medium", color="003366"),
)

_FONT_TITLE  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
_FONT_HEADER = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
_FONT_SUB    = Font(name="Calibri", bold=True, size=10, color="003366")
_FONT_TOTAL  = Font(name="Calibri", bold=True, size=10)
_FONT_NORMAL = Font(name="Calibri", size=10)
_FONT_WARN   = Font(name="Calibri", bold=True, size=10, color="CC0000")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
_ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")

_NUM_FMT = '#,##0.0'
_NUM_FMT_INT = '#,##0'

# Positionen, die als Subtotal / Zwischensumme formatiert werden
_SUBTOTALS_GUV = {
    "Gesamtumsatz", "Summe sonstige Erträge", "GESAMTLEISTUNG",
    "Summe Aufwendungen", "EBIT", "Summe Finanzergebnis",
    "EBT (Ergebnis vor Steuern)", "JAHRESÜBERSCHUSS", "Konzernergebnis",
}
_SUBTOTALS_BILANZ = {
    "Summe Anlagevermögen", "Summe Umlaufvermögen", "BILANZSUMME AKTIVA",
    "Summe Eigenkapital", "Summe Rückstellungen", "Summe Verbindlichkeiten",
    "BILANZSUMME PASSIVA",
}
_TOTALS = {"BILANZSUMME AKTIVA", "BILANZSUMME PASSIVA", "Konzernergebnis", "GESAMTLEISTUNG"}


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def _set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _title_row(ws, text: str, col_span: int, row: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _FONT_TITLE
    cell.fill = _HEADER
    cell.alignment = _ALIGN_CENTER


def _header_row(ws, labels: list[str], row: int, fill=_SUBHDR) -> None:
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = _FONT_HEADER
        c.fill = fill
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN


def _write_number(cell, value, is_subtotal: bool = False, is_total: bool = False):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        cell.value = None
        return
    cell.value = round(float(value), 1)
    cell.number_format = _NUM_FMT
    cell.alignment = _ALIGN_RIGHT
    cell.font = _FONT_TOTAL if (is_subtotal or is_total) else _FONT_NORMAL
    if is_total:
        cell.border = _BORDER_THICK


def _write_label(cell, text: str, indent: int = 0, is_subtotal: bool = False,
                 is_total: bool = False):
    cell.value = ("    " * indent) + str(text)
    cell.alignment = _ALIGN_LEFT
    cell.font = _FONT_TOTAL if (is_subtotal or is_total) else _FONT_NORMAL
    if is_total:
        cell.border = _BORDER_THICK


# ---------------------------------------------------------------------------
# Sheet 1 – Konzern-GuV
# ---------------------------------------------------------------------------

def _sheet_guv(wb: Workbook, guv: pd.DataFrame) -> None:
    ws = wb.create_sheet("Konzern-GuV")
    ws.freeze_panes = "B4"

    _title_row(ws, "KONZERN-GuV 2024  |  Muster Holding AG  |  Angaben in TEUR", 4, 1)
    _header_row(ws, ["Position", "2024 (TEUR)", "2023 (TEUR)", "Δ (%)"], 2)

    _set_col_width(ws, 1, 40)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 12)

    ws.row_dimensions[1].height = 24

    for r, (pos, row) in enumerate(guv.iterrows(), start=3):
        v24 = row["2024"]
        v23 = row["2023"]
        is_sub = pos in _SUBTOTALS_GUV
        is_tot = pos in _TOTALS

        fill = _GRAY if is_sub else _BLUE
        indent = 0 if is_sub else 1

        c_pos  = ws.cell(row=r, column=1)
        c_24   = ws.cell(row=r, column=2)
        c_23   = ws.cell(row=r, column=3)
        c_delt = ws.cell(row=r, column=4)

        _write_label(c_pos, pos, indent=indent, is_subtotal=is_sub, is_total=is_tot)
        _write_number(c_24, v24, is_subtotal=is_sub, is_total=is_tot)
        _write_number(c_23, v23, is_subtotal=is_sub, is_total=is_tot)

        c_pos.fill = fill
        c_24.fill  = fill
        c_23.fill  = fill

        # Delta %
        if v23 and abs(v23) > 0.1 and v24 is not None:
            delta = (v24 - v23) / abs(v23)
            c_delt.value = delta
            c_delt.number_format = "0.0%"
            c_delt.alignment = _ALIGN_RIGHT
            c_delt.font = _FONT_NORMAL
            if delta < -0.1:
                c_delt.font = Font(name="Calibri", size=10, color="CC0000")
        c_delt.fill = _GRAY


# ---------------------------------------------------------------------------
# Sheet 2 – Konzern-Bilanz
# ---------------------------------------------------------------------------

def _sheet_bilanz(wb: Workbook, bilanz: pd.DataFrame) -> None:
    ws = wb.create_sheet("Konzern-Bilanz")
    ws.freeze_panes = "A4"

    _title_row(ws, "KONZERN-BILANZ 31.12.2024  |  Muster Holding AG  |  Angaben in TEUR", 6, 1)
    _header_row(ws, ["AKTIVA", "2024", "2023", "PASSIVA", "2024", "2023"], 2)

    _set_col_width(ws, 1, 34)
    _set_col_width(ws, 2, 14)
    _set_col_width(ws, 3, 14)
    _set_col_width(ws, 4, 34)
    _set_col_width(ws, 5, 14)
    _set_col_width(ws, 6, 14)
    ws.row_dimensions[1].height = 24

    aktiva  = bilanz[bilanz["Seite"] == "Aktiva"].reset_index(drop=True)
    passiva = bilanz[bilanz["Seite"] == "Passiva"].reset_index(drop=True)
    n_rows  = max(len(aktiva), len(passiva))

    for i in range(n_rows):
        r = i + 3

        # Aktiva-Seite
        if i < len(aktiva):
            pos = aktiva.loc[i, "Position"]
            v24 = aktiva.loc[i, "2024"]
            v23 = aktiva.loc[i, "2023"]
            is_sub = pos in _SUBTOTALS_BILANZ
            is_tot = pos in _TOTALS
            fill = _GRAY if is_sub else _BLUE
            indent = 0 if is_sub else 1

            ca_pos = ws.cell(row=r, column=1)
            ca_24  = ws.cell(row=r, column=2)
            ca_23  = ws.cell(row=r, column=3)
            _write_label(ca_pos, pos, indent=indent, is_subtotal=is_sub, is_total=is_tot)
            _write_number(ca_24, v24, is_subtotal=is_sub, is_total=is_tot)
            _write_number(ca_23, v23, is_subtotal=is_sub, is_total=is_tot)
            ca_pos.fill = fill; ca_24.fill = fill; ca_23.fill = fill

        # Passiva-Seite
        if i < len(passiva):
            pos = passiva.loc[i, "Position"]
            v24 = passiva.loc[i, "2024"]
            v23 = passiva.loc[i, "2023"]
            is_sub = pos in _SUBTOTALS_BILANZ
            is_tot = pos in _TOTALS
            fill = _GRAY if is_sub else _BLUE
            indent = 0 if is_sub else 1

            cp_pos = ws.cell(row=r, column=4)
            cp_24  = ws.cell(row=r, column=5)
            cp_23  = ws.cell(row=r, column=6)
            _write_label(cp_pos, pos, indent=indent, is_subtotal=is_sub, is_total=is_tot)
            _write_number(cp_24, v24, is_subtotal=is_sub, is_total=is_tot)
            _write_number(cp_23, v23, is_subtotal=is_sub, is_total=is_tot)
            cp_pos.fill = fill; cp_24.fill = fill; cp_23.fill = fill

            # WAP und Konsolidierungsdifferenz in Rot
            if pos in ("Währungsausgleichsposten", "Konsolidierungsdifferenz"):
                for cell in (cp_pos, cp_24, cp_23):
                    cell.fill = _RED
                    cell.font = _FONT_WARN


# ---------------------------------------------------------------------------
# Sheet 3 – Kapitalflussrechnung (indirekte Methode)
# ---------------------------------------------------------------------------

def _sheet_kfr(wb: Workbook, guv: pd.DataFrame, bilanz: pd.DataFrame) -> None:
    ws = wb.create_sheet("Kapitalflussrechnung")
    ws.freeze_panes = "B4"

    _title_row(ws, "KAPITALFLUSSRECHNUNG 2024  |  Indirekte Methode  |  Angaben in TEUR", 3, 1)
    _header_row(ws, ["Position", "2024 (TEUR)", "Hinweis"], 2)
    _set_col_width(ws, 1, 44)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 30)
    ws.row_dimensions[1].height = 24

    def _get_guv(pos: str) -> float:
        return float(guv.loc[pos, "2024"]) if pos in guv.index else 0.0

    def _get_bilanz_delta(pos: str) -> float:
        row = bilanz[(bilanz["Position"] == pos) & (bilanz["Seite"].isin(["Aktiva", "Passiva"]))]
        if row.empty:
            return 0.0
        v24 = row["2024"].iloc[0]
        v23 = row["2023"].iloc[0]
        if pd.isna(v24) or pd.isna(v23):
            return 0.0
        return float(v24) - float(v23)

    jue    = _get_guv("JAHRESÜBERSCHUSS")
    afa    = -_get_guv("Abschreibungen (AfA)")      # AfA ist negativ in GuV → umkehren
    steuer = -_get_guv("Ertragsteuern (30%)")        # Steuer ist negativ → als Aufwand positiv

    delta_vorraete = _get_bilanz_delta("Vorräte")          # Anstieg = Mittelabfluss
    delta_ford     = _get_bilanz_delta("Forderungen L&L")  # Anstieg = Mittelabfluss
    delta_verbindl = _get_bilanz_delta("Verbindlichkeiten L&L")  # Anstieg = Mittelzufluss

    cfo = jue + afa - delta_vorraete - delta_ford + delta_verbindl
    # Näherung Investitionen: AfA als Proxy für Ersatzinvestitionen
    cfi = -afa * 1.1   # Kapex leicht über AfA (Wachstumsinvestitionen)
    # Finanzierung: Änderung Bankverbindlichkeiten
    delta_bank = _get_bilanz_delta("Bankverbindlichkeiten")
    delta_min  = _get_bilanz_delta("Minderheitenanteile")
    cff = delta_bank + delta_min - 0.0  # Dividenden nicht verfügbar
    netto = cfo + cfi + cff

    liquid_start = _get_bilanz_delta("Liquide Mittel")  # Δ
    liquid_end   = bilanz.loc[bilanz["Position"] == "Liquide Mittel", "2024"]
    liquid_end_v = float(liquid_end.iloc[0]) if not liquid_end.empty else 0.0

    rows_kfr = [
        # (label, wert, fill, hinweis)
        ("I. CASHFLOW AUS BETRIEBSTÄTIGKEIT", None, _SUBHDR, ""),
        ("  Jahresüberschuss",                jue,  _BLUE,   "aus Konzern-GuV"),
        ("  + Abschreibungen",                afa,  _BLUE,   "aus Konzern-GuV (nicht zahlungswirksam)"),
        ("  − Zunahme Vorräte",              -delta_vorraete, _BLUE, "Δ Bilanz 2024 vs. 2023"),
        ("  − Zunahme Forderungen L&L",      -delta_ford,     _BLUE, "Δ Bilanz 2024 vs. 2023"),
        ("  + Zunahme Verbindlichkeiten L&L", delta_verbindl, _BLUE, "Δ Bilanz 2024 vs. 2023"),
        ("  Cashflow aus Betriebstätigkeit",  cfo,  _GRAY,   "Summe I."),
        ("", None, None, ""),
        ("II. CASHFLOW AUS INVESTITIONEN",    None, _SUBHDR, ""),
        ("  − Investitionen (Näherung AfA)",  cfi,  _GREEN,  "Proxy: 110% der AfA"),
        ("  Cashflow aus Investitionen",      cfi,  _GRAY,   "Summe II."),
        ("", None, None, ""),
        ("III. CASHFLOW AUS FINANZIERUNG",    None, _SUBHDR, ""),
        ("  ± Veränderung Bankverbindl.",     delta_bank, _BLUE, "Δ Bilanz 2024 vs. 2023"),
        ("  + Veränderung Minderheitenanteile", delta_min, _BLUE, "Δ Bilanz 2024 vs. 2023"),
        ("  Cashflow aus Finanzierung",       cff,  _GRAY,   "Summe III. (ohne Dividenden)"),
        ("", None, None, ""),
        ("NETTO-VERÄNDERUNG LIQUIDE MITTEL",  netto, _GRAY,  "I. + II. + III."),
        ("  Prüfung: Δ Liquide Mittel Bilanz", liquid_start, _GREEN, "Bilanz 2024 minus 2023"),
        ("  Abweichung (Rundung / Proxy)",    round(netto - liquid_start, 1), _RED, "Soll ≈ 0"),
        ("", None, None, ""),
        ("Liquide Mittel 31.12.2024",         liquid_end_v, _BLUE, "aus Konzern-Bilanz"),
    ]

    for i, (label, wert, fill, hinweis) in enumerate(rows_kfr, start=3):
        is_header = fill is _SUBHDR
        is_total  = label.startswith("NETTO") or label.startswith("Cashflow")

        c_l = ws.cell(row=i, column=1, value=label)
        c_l.font = _FONT_HEADER if is_header else (_FONT_TOTAL if is_total else _FONT_NORMAL)
        c_l.alignment = _ALIGN_LEFT

        c_v = ws.cell(row=i, column=2)
        c_h = ws.cell(row=i, column=3, value=hinweis)
        c_h.font = Font(name="Calibri", size=9, italic=True, color="666666")

        if wert is not None:
            _write_number(c_v, wert, is_subtotal=is_total, is_total=(label.startswith("NETTO")))

        if fill:
            for cell in (c_l, c_v, c_h):
                cell.fill = fill
            if is_header:
                c_l.font = _FONT_HEADER
                c_l.fill = _SUBHDR


# ---------------------------------------------------------------------------
# Sheet 4 – Segmentbericht
# ---------------------------------------------------------------------------

def _sheet_segment(wb: Workbook, daten: dict) -> None:
    ws = wb.create_sheet("Segmentbericht")

    companies = list(daten.keys())
    headers = ["Kennzahl"] + [daten[k]["name"] for k in companies] + ["Konzern"]
    n_cols = len(headers)

    _title_row(ws, "SEGMENTBERICHT 2024  |  Angaben in TEUR", n_cols, 1)
    _header_row(ws, headers, 2)
    ws.row_dimensions[1].height = 24
    _set_col_width(ws, 1, 28)
    for col in range(2, n_cols + 2):
        _set_col_width(ws, col, 16)

    def _guv(key: str, pos: str) -> float:
        g = daten[key]["guv"]
        return float(g.loc[pos, "2024"]) if pos in g.index else 0.0

    def _bilanz(key: str, pos: str, seite: str) -> float:
        b = daten[key]["bilanz"]
        mask = (b["Position"] == pos) & (b["Seite"] == seite)
        if mask.any():
            return float(b.loc[mask, "2024"].iloc[0])
        return 0.0

    kennzahlen = [
        ("Umsatz",          lambda k: _guv(k, "Gesamtumsatz")),
        ("EBIT",            lambda k: _guv(k, "EBIT")),
        ("EBIT-Marge",      None),   # berechnet
        ("JAHRESÜBERSCHUSS",lambda k: _guv(k, "JAHRESÜBERSCHUSS")),
        ("Bilanzsumme",     lambda k: _bilanz(k, "BILANZSUMME AKTIVA", "Aktiva")),
        ("Eigenkapital",    lambda k: _bilanz(k, "Summe Eigenkapital", "Passiva")),
        ("Liquide Mittel",  lambda k: _bilanz(k, "Liquide Mittel", "Aktiva")),
        ("Mitarbeiter (n/a)", None),
        ("Land",            lambda k: daten[k].get("land", "")),
        ("Währung (orig.)", lambda k: daten[k].get("waehrung_orig", "EUR")),
        ("Beteiligung",     lambda k: f"{daten[k].get('beteiligung', 1.0):.0%}"),
    ]

    for r_off, (label, fn) in enumerate(kennzahlen, start=3):
        ws.cell(row=r_off, column=1, value=label).font = _FONT_SUB

        row_vals: list = []
        for col, key in enumerate(companies, start=2):
            cell = ws.cell(row=r_off, column=col)
            cell.fill = _BLUE

            if label == "EBIT-Marge":
                umsatz = _guv(key, "Gesamtumsatz")
                ebit   = _guv(key, "EBIT")
                val = ebit / umsatz if umsatz else 0.0
                cell.value = val
                cell.number_format = "0.0%"
                cell.alignment = _ALIGN_RIGHT
                row_vals.append(val)
            elif label == "Mitarbeiter (n/a)":
                cell.value = "n/v"
                cell.alignment = _ALIGN_CENTER
                row_vals.append(None)
            elif fn is not None:
                val = fn(key)
                if isinstance(val, (int, float)):
                    _write_number(cell, val)
                    row_vals.append(val)
                else:
                    cell.value = val
                    cell.alignment = _ALIGN_CENTER
                    row_vals.append(None)
            else:
                row_vals.append(None)

        # Konzern-Spalte
        konzern_cell = ws.cell(row=r_off, column=len(companies) + 2)
        konzern_cell.fill = _GREEN
        if label in ("Land", "Währung (orig.)", "Beteiligung", "Mitarbeiter (n/a)"):
            konzern_cell.value = "Konzern"
            konzern_cell.alignment = _ALIGN_CENTER
        elif label == "EBIT-Marge":
            sum_umsatz = sum(_guv(k, "Gesamtumsatz") for k in companies)
            sum_ebit   = sum(_guv(k, "EBIT") for k in companies)
            konzern_cell.value = sum_ebit / sum_umsatz if sum_umsatz else 0
            konzern_cell.number_format = "0.0%"
            konzern_cell.alignment = _ALIGN_RIGHT
        elif fn is not None and all(isinstance(v, (int, float)) for v in row_vals if v is not None):
            _write_number(konzern_cell, sum(v for v in row_vals if v is not None), is_total=True)


# ---------------------------------------------------------------------------
# Sheet 5 – IC-Eliminierungen
# ---------------------------------------------------------------------------

def _sheet_ic(wb: Workbook, ic_log: dict) -> None:
    ws = wb.create_sheet("IC-Eliminierungen")

    _title_row(ws, "IC-ELIMINIERUNGEN – NACHWEIS  |  Angaben in TEUR", 5, 1)
    ws.row_dimensions[1].height = 24
    _set_col_width(ws, 1, 34)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 16)
    _set_col_width(ws, 5, 28)

    # --- GuV-Teil ---
    ws.cell(row=2, column=1, value="I. AUFWANDS-/ERTRAGSKONSOLIDIERUNG (GuV)").font = _FONT_SUB
    _header_row(ws, ["Gesellschaft", "IC-Ertrag", "IC-Aufwand", "Netto", "Positionen"], 3,
                fill=_SUBHDR)

    g = ic_log["guv"]
    row = 4
    for key, d in g["detail"].items():
        if d["ic_ertrag_elim"] == 0 and d["ic_aufwand_elim"] == 0:
            continue
        ws.cell(row=row, column=1, value=key).fill = _BLUE
        _write_number(ws.cell(row=row, column=2), d["ic_ertrag_elim"])
        _write_number(ws.cell(row=row, column=3), d["ic_aufwand_elim"])
        _write_number(ws.cell(row=row, column=4), d["ic_ertrag_elim"] + d["ic_aufwand_elim"])
        pos_str = "; ".join(d["positionen_ertrag"] + d["positionen_aufwand"])[:60]
        ws.cell(row=row, column=5, value=pos_str).font = Font(name="Calibri", size=9, italic=True)
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = _BLUE
        row += 1

    # Summenzeile
    ws.cell(row=row, column=1, value="SUMME GuV").font = _FONT_TOTAL
    _write_number(ws.cell(row=row, column=2), g["total_ertrag"], is_total=True)
    _write_number(ws.cell(row=row, column=3), g["total_aufwand"], is_total=True)
    saldo_fill = _GRAY if abs(g["saldo"]) <= 0.5 else _RED
    saldo_cell = ws.cell(row=row, column=4)
    _write_number(saldo_cell, g["saldo"], is_total=True)
    saldo_cell.fill = saldo_fill
    ws.cell(row=row, column=5, value="✅ ausgeglichen" if abs(g["saldo"]) <= 0.5
            else f"⚠️  Saldo {g['saldo']:+.1f} TEUR (Zwischengewinne)").fill = saldo_fill
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = _GRAY
    row += 2

    # --- Bilanz-Teil ---
    ws.cell(row=row, column=1, value="II. SCHULDENKONSOLIDIERUNG (Bilanz)").font = _FONT_SUB
    row += 1
    _header_row(ws, ["Gesellschaft", "IC-Forderung", "IC-Verbindl.", "Saldo", ""], row,
                fill=_SUBHDR)
    row += 1

    b = ic_log["bilanz"]
    for key, d in b["detail"].items():
        if d["ic_forderung_elim"] == 0 and d["ic_verbindlichkeit_elim"] == 0:
            continue
        ws.cell(row=row, column=1, value=key).fill = _BLUE
        _write_number(ws.cell(row=row, column=2), d["ic_forderung_elim"])
        _write_number(ws.cell(row=row, column=3), d["ic_verbindlichkeit_elim"])
        _write_number(ws.cell(row=row, column=4),
                      d["ic_forderung_elim"] - d["ic_verbindlichkeit_elim"])
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = _BLUE
        row += 1

    ws.cell(row=row, column=1, value="SUMME Bilanz").font = _FONT_TOTAL
    _write_number(ws.cell(row=row, column=2), b["total_forderung"], is_total=True)
    _write_number(ws.cell(row=row, column=3), b["total_verbindlichkeit"], is_total=True)
    saldo_fill = _GRAY if abs(b["saldo"]) <= 0.5 else _RED
    saldo_cell = ws.cell(row=row, column=4)
    _write_number(saldo_cell, b["saldo"], is_total=True)
    saldo_cell.fill = saldo_fill
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = _GRAY


# ---------------------------------------------------------------------------
# Öffentliche API
# ---------------------------------------------------------------------------

def exportiere_excel(
    guv: pd.DataFrame,
    bilanz: pd.DataFrame,
    daten: dict,
    ic_log: dict,
    pfad: Path | None = None,
) -> Path:
    """
    Erstellt den Konzernabschluss als .xlsx.

    Rückgabe: Pfad zur erstellten Datei.
    """
    if pfad is None:
        pfad = OUTPUT_DIR / "Konzernabschluss_2024.xlsx"

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)   # Standard-Sheet entfernen

    log.info("Sheet 1/5: Konzern-GuV …")
    _sheet_guv(wb, guv)

    log.info("Sheet 2/5: Konzern-Bilanz …")
    _sheet_bilanz(wb, bilanz)

    log.info("Sheet 3/5: Kapitalflussrechnung …")
    _sheet_kfr(wb, guv, bilanz)

    log.info("Sheet 4/5: Segmentbericht …")
    _sheet_segment(wb, daten)

    log.info("Sheet 5/5: IC-Eliminierungen …")
    _sheet_ic(wb, ic_log)

    wb.save(str(pfad))
    log.info("✅ Gespeichert: %s  (%.1f KB)", pfad, pfad.stat().st_size / 1024)
    return pfad


# ---------------------------------------------------------------------------
# Standalone-Ausführung
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import importlib.util

    def _lm(name: str, p: Path):
        s = importlib.util.spec_from_file_location(name, p)
        m = importlib.util.module_from_spec(s); s.loader.exec_module(m); return m

    src = ROOT / "src"
    m1 = _lm("load_data", src / "01_load_data.py")
    m2 = _lm("fx",        src / "02_fx_conversion.py")
    m3 = _lm("ic",        src / "03_ic_elimination.py")
    m4 = _lm("mi",        src / "04_minority_interest.py")
    m5 = _lm("co",        src / "05_consolidate.py")

    daten          = m1.load_all()
    kurse          = m2.lade_fx_kurse()
    daten          = m2.konvertiere_alle(daten, kurse)
    daten, ic_log  = m3.eliminiere_ic(daten)
    daten, _       = m4.berechne_alle_minderheiten(daten)
    guv, bilanz    = m5.konsolidiere(daten)

    exportiere_excel(guv, bilanz, daten, ic_log)
